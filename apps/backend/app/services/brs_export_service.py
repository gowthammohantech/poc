import csv
import io
from typing import Any, Dict


def build_brs_export_json(final_output: Dict[str, Any]) -> Dict[str, Any]:
    return final_output.get("corrected_json", {})


def build_brs_export_csv(final_output: Dict[str, Any]) -> str:
    corrected = final_output.get("corrected_json", {})
    brs = corrected.get("brs", corrected)
    doc_info = brs.get("document_info", {})
    balances = brs.get("balances", {})
    bank_items = brs.get("bank_side_items", [])
    book_items = brs.get("book_side_items", [])

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(["BRS Document Info"])
    writer.writerow(["Company Name", doc_info.get("company_name", "")])
    writer.writerow(["Bank Name", doc_info.get("bank_name", "")])
    writer.writerow(["Account Number", doc_info.get("account_number", "")])
    writer.writerow(["Statement Period Start", doc_info.get("statement_period_start", "")])
    writer.writerow(["Statement Period End", doc_info.get("statement_period_end", "")])
    writer.writerow(["Currency", doc_info.get("currency", "")])
    writer.writerow(["Prepared By", doc_info.get("prepared_by", "")])
    writer.writerow(["Prepared Date", doc_info.get("prepared_date", "")])
    writer.writerow([])

    writer.writerow(["Balances"])
    writer.writerow(["Opening Balance (Bank)", balances.get("opening_balance_bank", "")])
    writer.writerow(["Opening Balance (Book)", balances.get("opening_balance_book", "")])
    writer.writerow(["Closing Balance (Bank)", balances.get("closing_balance_bank", "")])
    writer.writerow(["Closing Balance (Book)", balances.get("closing_balance_book", "")])
    writer.writerow(["Reconciled Balance", balances.get("reconciled_balance", "")])
    writer.writerow(["Adjusted Bank Balance", brs.get("adjusted_bank_balance", "")])
    writer.writerow(["Adjusted Book Balance", brs.get("adjusted_book_balance", "")])
    writer.writerow([])

    writer.writerow(["Bank Side Items"])
    writer.writerow(["Type", "Description", "Reference #", "Date", "Amount", "Effect"])
    for item in bank_items:
        writer.writerow([
            item.get("item_type", ""),
            item.get("description", ""),
            item.get("reference_number", ""),
            item.get("date", ""),
            item.get("amount", ""),
            item.get("effect", ""),
        ])
    writer.writerow([])

    writer.writerow(["Book Side Items"])
    writer.writerow(["Type", "Description", "Reference #", "Date", "Amount", "Effect"])
    for item in book_items:
        writer.writerow([
            item.get("item_type", ""),
            item.get("description", ""),
            item.get("reference_number", ""),
            item.get("date", ""),
            item.get("amount", ""),
            item.get("effect", ""),
        ])

    return output.getvalue()


def build_brs_export_excel(final_output: Dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    corrected = final_output.get("corrected_json", {})
    brs = corrected.get("brs", corrected)
    doc_info = brs.get("document_info", {})
    balances = brs.get("balances", {})
    bank_items = brs.get("bank_side_items", [])
    book_items = brs.get("book_side_items", [])

    wb = Workbook()
    bold = Font(bold=True)

    # Sheet 1: BRS Header
    ws1 = wb.active
    ws1.title = "BRS Header"
    header_data = [
        ("Company Name", doc_info.get("company_name")),
        ("Bank Name", doc_info.get("bank_name")),
        ("Account Number", doc_info.get("account_number")),
        ("Period Start", doc_info.get("statement_period_start")),
        ("Period End", doc_info.get("statement_period_end")),
        ("Currency", doc_info.get("currency")),
        ("Prepared By", doc_info.get("prepared_by")),
        ("Prepared Date", doc_info.get("prepared_date")),
        ("", ""),
        ("Opening Balance (Bank)", balances.get("opening_balance_bank")),
        ("Opening Balance (Book)", balances.get("opening_balance_book")),
        ("Closing Balance (Bank)", balances.get("closing_balance_bank")),
        ("Closing Balance (Book)", balances.get("closing_balance_book")),
        ("Reconciled Balance", balances.get("reconciled_balance")),
        ("Adjusted Bank Balance", brs.get("adjusted_bank_balance")),
        ("Adjusted Book Balance", brs.get("adjusted_book_balance")),
    ]
    for row_idx, (label, value) in enumerate(header_data, start=1):
        ws1.cell(row=row_idx, column=1, value=label).font = bold
        ws1.cell(row=row_idx, column=2, value=value)
    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 35

    # Sheet 2: Bank Side Items
    item_columns = ["Type", "Description", "Reference #", "Date", "Amount", "Effect"]
    for sheet_title, items in [("Bank Side Items", bank_items), ("Book Side Items", book_items)]:
        ws = wb.create_sheet(sheet_title)
        for col_idx, col_name in enumerate(item_columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = bold
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
        for row_idx, item in enumerate(items, start=2):
            ws.cell(row=row_idx, column=1, value=item.get("item_type"))
            ws.cell(row=row_idx, column=2, value=item.get("description"))
            ws.cell(row=row_idx, column=3, value=item.get("reference_number"))
            ws.cell(row=row_idx, column=4, value=item.get("date"))
            ws.cell(row=row_idx, column=5, value=item.get("amount"))
            ws.cell(row=row_idx, column=6, value=item.get("effect"))
        for col_idx in range(1, len(item_columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
