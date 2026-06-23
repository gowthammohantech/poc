import csv
import io
import json
from typing import Any, Dict


def build_export_json(final_output: Dict[str, Any], document: Dict[str, Any]) -> Dict[str, Any]:
    corrected = final_output.get("corrected_json", {})
    return corrected


def build_export_csv(final_output: Dict[str, Any]) -> str:
    corrected = final_output.get("corrected_json", {})
    invoice = corrected.get("invoice", {})
    line_items = invoice.get("line_items", [])

    output = io.StringIO()
    writer = csv.writer(output)

    # Header info
    writer.writerow(["Invoice Number", invoice.get("invoice_number", "")])
    writer.writerow(["Invoice Date", invoice.get("invoice_date", "")])
    writer.writerow(["Due Date", invoice.get("due_date", "")])
    writer.writerow(["Currency", invoice.get("currency", "")])
    writer.writerow([])

    writer.writerow(["Vendor Name", invoice.get("vendor", {}).get("name", "")])
    writer.writerow(["Vendor GSTIN", invoice.get("vendor", {}).get("gstin", "")])
    writer.writerow([])

    writer.writerow(["Customer Name", invoice.get("customer", {}).get("name", "")])
    writer.writerow(["Customer GSTIN", invoice.get("customer", {}).get("gstin", "")])
    writer.writerow([])

    # Line items
    writer.writerow(["Line Items"])
    writer.writerow([
        "Description", "HSN/SAC", "Quantity", "Unit", "Unit Price",
        "Taxable Value", "CGST Rate", "CGST Amt", "SGST Rate", "SGST Amt",
        "IGST Rate", "IGST Amt", "Total"
    ])
    for item in line_items:
        writer.writerow([
            item.get("description", ""),
            item.get("hsn_sac", ""),
            item.get("quantity", ""),
            item.get("unit", ""),
            item.get("unit_price", ""),
            item.get("taxable_value", ""),
            item.get("cgst_rate", ""),
            item.get("cgst_amount", ""),
            item.get("sgst_rate", ""),
            item.get("sgst_amount", ""),
            item.get("igst_rate", ""),
            item.get("igst_amount", ""),
            item.get("total", ""),
        ])
    writer.writerow([])

    # Tax summary
    tax = invoice.get("tax_summary", {})
    writer.writerow(["Tax Summary"])
    writer.writerow(["Subtotal", tax.get("subtotal", "")])
    writer.writerow(["CGST Total", tax.get("cgst_total", "")])
    writer.writerow(["SGST Total", tax.get("sgst_total", "")])
    writer.writerow(["IGST Total", tax.get("igst_total", "")])
    writer.writerow(["Total Tax", tax.get("total_tax", "")])
    writer.writerow(["Round Off", tax.get("round_off", "")])
    writer.writerow(["Grand Total", tax.get("grand_total", "")])

    return output.getvalue()


def build_export_excel(final_output: Dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    corrected = final_output.get("corrected_json", {})
    invoice = corrected.get("invoice", {})
    line_items = invoice.get("line_items", [])
    tax = invoice.get("tax_summary", {})

    wb = Workbook()

    # Sheet 1: Invoice Header
    ws1 = wb.active
    ws1.title = "Invoice Header"
    header_style = Font(bold=True)

    header_data = [
        ("Invoice Number", invoice.get("invoice_number")),
        ("Invoice Date", invoice.get("invoice_date")),
        ("Due Date", invoice.get("due_date")),
        ("Currency", invoice.get("currency")),
        ("Purchase Order", invoice.get("purchase_order_number")),
        ("", ""),
        ("Vendor Name", invoice.get("vendor", {}).get("name")),
        ("Vendor GSTIN", invoice.get("vendor", {}).get("gstin")),
        ("Vendor PAN", invoice.get("vendor", {}).get("pan")),
        ("Vendor Address", invoice.get("vendor", {}).get("address")),
        ("Vendor Phone", invoice.get("vendor", {}).get("phone")),
        ("Vendor Email", invoice.get("vendor", {}).get("email")),
        ("", ""),
        ("Customer Name", invoice.get("customer", {}).get("name")),
        ("Customer GSTIN", invoice.get("customer", {}).get("gstin")),
        ("Customer Address", invoice.get("customer", {}).get("address")),
        ("", ""),
        ("Subtotal", tax.get("subtotal")),
        ("CGST Total", tax.get("cgst_total")),
        ("SGST Total", tax.get("sgst_total")),
        ("IGST Total", tax.get("igst_total")),
        ("Total Tax", tax.get("total_tax")),
        ("Round Off", tax.get("round_off")),
        ("Grand Total", tax.get("grand_total")),
    ]
    for row_idx, (label, value) in enumerate(header_data, start=1):
        ws1.cell(row=row_idx, column=1, value=label).font = header_style
        ws1.cell(row=row_idx, column=2, value=value)

    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 40

    # Sheet 2: Line Items
    ws2 = wb.create_sheet("Line Items")
    columns = [
        "Description", "HSN/SAC", "Quantity", "Unit", "Unit Price",
        "Taxable Value", "CGST Rate%", "CGST Amount", "SGST Rate%", "SGST Amount",
        "IGST Rate%", "IGST Amount", "Total"
    ]
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    for row_idx, item in enumerate(line_items, start=2):
        ws2.cell(row=row_idx, column=1, value=item.get("description"))
        ws2.cell(row=row_idx, column=2, value=item.get("hsn_sac"))
        ws2.cell(row=row_idx, column=3, value=item.get("quantity"))
        ws2.cell(row=row_idx, column=4, value=item.get("unit"))
        ws2.cell(row=row_idx, column=5, value=item.get("unit_price"))
        ws2.cell(row=row_idx, column=6, value=item.get("taxable_value"))
        ws2.cell(row=row_idx, column=7, value=item.get("cgst_rate"))
        ws2.cell(row=row_idx, column=8, value=item.get("cgst_amount"))
        ws2.cell(row=row_idx, column=9, value=item.get("sgst_rate"))
        ws2.cell(row=row_idx, column=10, value=item.get("sgst_amount"))
        ws2.cell(row=row_idx, column=11, value=item.get("igst_rate"))
        ws2.cell(row=row_idx, column=12, value=item.get("igst_amount"))
        ws2.cell(row=row_idx, column=13, value=item.get("total"))

    for col_idx in range(1, len(columns) + 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
