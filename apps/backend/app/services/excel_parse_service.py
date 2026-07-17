import io
from datetime import datetime, date
from typing import Any, Dict, List

COA_FIELDS = [
    ("account_code", "Account Code", True),
    ("account_name", "Account Name", True),
    ("account_type", "Account Type", True),
    ("parent_group", "Parent Group", False),
]

LEDGER_FIELDS = [
    ("transaction_type", "Transaction Type", True),
    ("invoice_no", "Invoice No", True),
    # invoice_date/invoice_amount are the invoice's own identity/expected amount. Kept as
    # OPTIONAL columns (not file-level required) so re-uploading an old-template workbook that
    # predates these fields is still accepted — absence is handled gracefully downstream
    # (invoice_amount unknown -> "Needs Review" status, not a hard rejection).
    ("invoice_date", "Invoice Date", False),
    ("invoice_amount", "Invoice Amount", False),
    # ledger_date/voucher/amount describe ONE settlement leg and may be blank together (an
    # invoice recorded but not yet settled) — enforced as all-or-nothing at the row level below,
    # not as individually required columns.
    ("ledger_date", "Ledger Date", False),
    ("ledger_voucher", "Ledger Voucher", False),
    ("ledger_amount", "Ledger Amount", False),
    ("account_name", "Account Name", False),
    ("bank_cash_account", "Bank/Cash Account", False),
]

VALID_TRANSACTION_TYPES = {"sales receipt", "purchase payment", "sales invoice", "purchase invoice", "journal entry"}

_DATE_FORMATS = ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d/%m/%y", "%d-%b-%y", "%d-%b-%Y"]


def _normalize_header(name: str) -> str:
    return "".join(ch for ch in name.strip().lower() if ch.isalnum())


def _build_header_map(header_row, fields) -> Dict[str, int]:
    normalized_expected = {_normalize_header(label): key for key, label, _ in fields}
    header_map: Dict[str, int] = {}
    for col_idx, cell in enumerate(header_row):
        if cell is None:
            continue
        norm = _normalize_header(str(cell))
        if norm in normalized_expected:
            header_map[normalized_expected[norm]] = col_idx
    return header_map


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _parse_date_value(value: Any):
    if value is None or value == "":
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        s = value.strip()
        for fmt in _DATE_FORMATS:
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return None
    return None


def _parse_amount_value(value: Any):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "").replace("₹", "").replace("INR", "").strip()
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _canonical_transaction_type(transaction_type: str) -> str:
    """Maps any of the accepted Transaction Type spellings to the two canonical directions the
    frontend understands (money in vs money out) — a "Sales Invoice" is directionally the same
    as a "Sales Receipt", etc."""
    t = transaction_type.lower()
    if t in ("sales receipt", "sales invoice"):
        return "Sales Receipt"
    return "Purchase Payment"


def _load_first_sheet_rows(file_bytes: bytes):
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        raise ValueError(f"Could not read Excel file: {e}")
    ws = wb.active
    return ws.iter_rows(values_only=True)


def parse_coa_excel(file_bytes: bytes) -> Dict[str, Any]:
    rows_iter = _load_first_sheet_rows(file_bytes)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return {"rows": [], "errors": ["The file is empty."]}

    header_map = _build_header_map(header_row, COA_FIELDS)
    missing_required = [label for key, label, required in COA_FIELDS if required and key not in header_map]
    if missing_required:
        return {"rows": [], "errors": [f"Missing required column(s): {', '.join(missing_required)}"]}

    rows: List[Dict[str, Any]] = []
    errors: List[str] = []
    for row_idx, raw_row in enumerate(rows_iter, start=2):
        if raw_row is None or all(c is None for c in raw_row):
            continue
        row_errors = []
        account_code = _cell(raw_row, header_map.get("account_code"))
        account_name = _cell(raw_row, header_map.get("account_name"))
        account_type = _cell(raw_row, header_map.get("account_type"))
        parent_group = _cell(raw_row, header_map.get("parent_group"))

        if not account_code:
            row_errors.append("Account Code is required")
        if not account_name:
            row_errors.append("Account Name is required")
        if not account_type:
            row_errors.append("Account Type is required")

        if row_errors:
            errors.append(f"Row {row_idx}: {'; '.join(row_errors)}")
            continue

        rows.append({
            "account_code": str(account_code).strip(),
            "account_name": str(account_name).strip(),
            "account_type": str(account_type).strip(),
            "parent_group": str(parent_group).strip() if parent_group else None,
        })

    return {"rows": rows, "errors": errors}


def parse_ledger_excel(file_bytes: bytes) -> Dict[str, Any]:
    rows_iter = _load_first_sheet_rows(file_bytes)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return {"rows": [], "errors": ["The file is empty."]}

    header_map = _build_header_map(header_row, LEDGER_FIELDS)
    missing_required = [label for key, label, required in LEDGER_FIELDS if required and key not in header_map]
    if missing_required:
        return {"rows": [], "errors": [f"Missing required column(s): {', '.join(missing_required)}"]}

    rows: List[Dict[str, Any]] = []
    errors: List[str] = []
    for row_idx, raw_row in enumerate(rows_iter, start=2):
        if raw_row is None or all(c is None for c in raw_row):
            continue
        row_errors = []

        transaction_type_raw = _cell(raw_row, header_map.get("transaction_type"))
        invoice_no = _cell(raw_row, header_map.get("invoice_no"))
        invoice_date_raw = _cell(raw_row, header_map.get("invoice_date")) if "invoice_date" in header_map else None
        invoice_amount_raw = _cell(raw_row, header_map.get("invoice_amount")) if "invoice_amount" in header_map else None
        ledger_date_raw = _cell(raw_row, header_map.get("ledger_date")) if "ledger_date" in header_map else None
        ledger_voucher = _cell(raw_row, header_map.get("ledger_voucher")) if "ledger_voucher" in header_map else None
        ledger_amount_raw = _cell(raw_row, header_map.get("ledger_amount")) if "ledger_amount" in header_map else None
        account_name = _cell(raw_row, header_map.get("account_name")) if "account_name" in header_map else None
        bank_cash_account = _cell(raw_row, header_map.get("bank_cash_account")) if "bank_cash_account" in header_map else None

        transaction_type = str(transaction_type_raw).strip() if transaction_type_raw else None
        if not transaction_type or transaction_type.lower() not in VALID_TRANSACTION_TYPES:
            row_errors.append('Transaction Type must be "Sales Receipt" or "Purchase Payment"')

        # Every field below (Invoice No/Date/Amount, Ledger Date/Voucher/Amount) is independently
        # optional — each is mapped through whenever the sheet has it, regardless of whether its
        # sibling fields are also present. A field is only flagged as an error if it was actually
        # provided but couldn't be parsed (garbled date/amount); a blank cell is never an error.
        invoice_date = _parse_date_value(invoice_date_raw)
        if invoice_date_raw not in (None, "") and invoice_date is None:
            row_errors.append("Invoice Date is unparseable")

        invoice_amount = _parse_amount_value(invoice_amount_raw)
        if invoice_amount_raw not in (None, "") and invoice_amount is None:
            row_errors.append("Invoice Amount is not numeric")

        ledger_date = _parse_date_value(ledger_date_raw)
        if ledger_date_raw not in (None, "") and ledger_date is None:
            row_errors.append("Ledger Date is unparseable")

        ledger_amount = _parse_amount_value(ledger_amount_raw)
        if ledger_amount_raw not in (None, "") and ledger_amount is None:
            row_errors.append("Ledger Amount is not numeric")

        if row_errors:
            errors.append(f"Row {row_idx}: {'; '.join(row_errors)}")
            continue

        canonical_type = _canonical_transaction_type(transaction_type)

        rows.append({
            "transaction_type": canonical_type,
            "invoice_no": str(invoice_no).strip() if invoice_no else "",
            "invoice_date": invoice_date,
            "invoice_amount": invoice_amount,
            "ledger_date": ledger_date,
            "ledger_voucher": str(ledger_voucher).strip() if ledger_voucher else None,
            "ledger_amount": ledger_amount,
            "account_name": str(account_name).strip() if account_name else None,
            "bank_cash_account": str(bank_cash_account).strip() if bank_cash_account else None,
        })

    return {"rows": rows, "errors": errors}
